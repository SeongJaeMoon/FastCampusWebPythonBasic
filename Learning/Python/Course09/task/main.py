from openpyxl import load_workbook as load
DIR = 'test.xlsx'

def save_excel(ws):
    idx = 65
    for i in range(1, 10):
        for j in range(1, 10):
            ws[chr(idx) + str(j)] = "{} x {} = {}".format(i, j, i * j)
        idx += 1
    wb.save(DIR)

'''
test.xlsx 저장 결과
    A           B           C           D           E           F           G           H           I
1 x 1 = 1	2 x 1 = 2	3 x 1 = 3	4 x 1 = 4	5 x 1 = 5	6 x 1 = 6	7 x 1 = 7	8 x 1 = 8	9 x 1 = 9
1 x 2 = 2	2 x 2 = 4	3 x 2 = 6	4 x 2 = 8	5 x 2 = 10	6 x 2 = 12	7 x 2 = 14	8 x 2 = 16	9 x 2 = 18
1 x 3 = 3	2 x 3 = 6	3 x 3 = 9	4 x 3 = 12	5 x 3 = 15	6 x 3 = 18	7 x 3 = 21	8 x 3 = 24	9 x 3 = 27
1 x 4 = 4	2 x 4 = 8	3 x 4 = 12	4 x 4 = 16	5 x 4 = 20	6 x 4 = 24	7 x 4 = 28	8 x 4 = 32	9 x 4 = 36
1 x 5 = 5	2 x 5 = 10	3 x 5 = 15	4 x 5 = 20	5 x 5 = 25	6 x 5 = 30	7 x 5 = 35	8 x 5 = 40	9 x 5 = 45
1 x 6 = 6	2 x 6 = 12	3 x 6 = 18	4 x 6 = 24	5 x 6 = 30	6 x 6 = 36	7 x 6 = 42	8 x 6 = 48	9 x 6 = 54
1 x 7 = 7	2 x 7 = 14	3 x 7 = 21	4 x 7 = 28	5 x 7 = 35	6 x 7 = 42	7 x 7 = 49	8 x 7 = 56	9 x 7 = 63
1 x 8 = 8	2 x 8 = 16	3 x 8 = 24	4 x 8 = 32	5 x 8 = 40	6 x 8 = 48	7 x 8 = 56	8 x 8 = 64	9 x 8 = 72
1 x 9 = 9	2 x 9 = 18	3 x 9 = 27	4 x 9 = 36	5 x 9 = 45	6 x 9 = 54	7 x 9 = 63	8 x 9 = 72	9 x 9 = 81

'''

def get_excel(ws):
    idx = 64
    for i in range(1, 10):
        for j in range(1, 10):
            print(ws[chr(idx + j) + str(i)].value, end = ' ')
        print()
        

'''
get_excel() 실행 결과
1 x 1 = 1 2 x 1 = 2 3 x 1 = 3 4 x 1 = 4 5 x 1 = 5 6 x 1 = 6 7 x 1 = 7 8 x 1 = 8 9 x 1 = 9
1 x 2 = 2 2 x 2 = 4 3 x 2 = 6 4 x 2 = 8 5 x 2 = 10 6 x 2 = 12 7 x 2 = 14 8 x 2 = 16 9 x 2 = 18
1 x 3 = 3 2 x 3 = 6 3 x 3 = 9 4 x 3 = 12 5 x 3 = 15 6 x 3 = 18 7 x 3 = 21 8 x 3 = 24 9 x 3 = 27
1 x 4 = 4 2 x 4 = 8 3 x 4 = 12 4 x 4 = 16 5 x 4 = 20 6 x 4 = 24 7 x 4 = 28 8 x 4 = 32 9 x 4 = 36
1 x 5 = 5 2 x 5 = 10 3 x 5 = 15 4 x 5 = 20 5 x 5 = 25 6 x 5 = 30 7 x 5 = 35 8 x 5 = 40 9 x 5 = 45
1 x 6 = 6 2 x 6 = 12 3 x 6 = 18 4 x 6 = 24 5 x 6 = 30 6 x 6 = 36 7 x 6 = 42 8 x 6 = 48 9 x 6 = 54
1 x 7 = 7 2 x 7 = 14 3 x 7 = 21 4 x 7 = 28 5 x 7 = 35 6 x 7 = 42 7 x 7 = 49 8 x 7 = 56 9 x 7 = 63
1 x 8 = 8 2 x 8 = 16 3 x 8 = 24 4 x 8 = 32 5 x 8 = 40 6 x 8 = 48 7 x 8 = 56 8 x 8 = 64 9 x 8 = 72
1 x 9 = 9 2 x 9 = 18 3 x 9 = 27 4 x 9 = 36 5 x 9 = 45 6 x 9 = 54 7 x 9 = 63 8 x 9 = 72 9 x 9 = 81
'''

if __name__ == "__main__":
    wb = load(DIR)
    ws = wb.create_sheet('99단')
    try:
        save_excel(ws)
        get_excel(ws)
    except Exception as e:
        print(e)
    finally:
        wb.close()    


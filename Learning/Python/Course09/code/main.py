from openpyxl import load_workbook as load

DIR = '/Users/moonseongjae/FastCampusPythonBasic/code09/test.xlsx'

wb = load(DIR)


# ws = wb.create_sheet('test') # 같은 이름의 시트 이름이 존재하면, 새로운 시트를 생성합니다.
# ws = wb['test']
ws['A1'] = "제목1"
ws['B1'] = "제목2"
a1 = ws['A2'].value
a2 = ws['B2'].value
print(a1, a2)
# wb.save(DIR)
wb.close()

